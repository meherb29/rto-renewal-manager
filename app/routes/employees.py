from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from app import db
from app.models import Employee
from flask import send_file
import openpyxl
from openpyxl.styles import Font
import io
from app.models import Renewal
from datetime import datetime

employees = Blueprint('employees', __name__)

@employees.route('/export')
@login_required
def export_data():
    if current_user.role != 'admin':
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard.index'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Renewals'

    headers = [
        'Customer Name', 'Phone', 'Branch', 'Is Company',
        'GST Number', 'PAN Number',
        'Vehicle Model', 'Category', 'Registration No',
        'Chassis No', 'Engine No', 'Passing Type', 'Year of Manufacture',
        'Insurer', 'Policy Number', 'Policy Start', 'Policy End',
        'IDV', 'Premium', 'Status', 'Reminder Due',
        'Data Confirmed', 'Last Known Year', 'Notes',
        'Assigned Employee'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    renewals = Renewal.query.order_by(Renewal.policy_end).all()

    for row, renewal in enumerate(renewals, 2):
        customer = renewal.vehicle.customer
        vehicle = renewal.vehicle
        ws.cell(row=row, column=1, value=customer.name)
        ws.cell(row=row, column=2, value=customer.phone)
        ws.cell(row=row, column=3, value=customer.branch)
        ws.cell(row=row, column=4, value='Yes' if customer.is_company else 'No')
        ws.cell(row=row, column=5, value=customer.gst_number)
        ws.cell(row=row, column=6, value=customer.pan_number)
        ws.cell(row=row, column=7, value=vehicle.model)
        ws.cell(row=row, column=8, value=vehicle.category)
        ws.cell(row=row, column=9, value=vehicle.registration_no)
        ws.cell(row=row, column=10, value=vehicle.chassis_no)
        ws.cell(row=row, column=11, value=vehicle.engine_no)
        ws.cell(row=row, column=12, value=vehicle.passing_type)
        ws.cell(row=row, column=13, value=vehicle.year_of_manufacture)
        ws.cell(row=row, column=14, value=renewal.insurer.name if renewal.insurer else None)
        ws.cell(row=row, column=15, value=renewal.policy_number)
        ws.cell(row=row, column=16, value=str(renewal.policy_start) if renewal.policy_start else None)
        ws.cell(row=row, column=17, value=str(renewal.policy_end) if renewal.policy_end else None)
        ws.cell(row=row, column=18, value=renewal.idv)
        ws.cell(row=row, column=19, value=renewal.premium)
        ws.cell(row=row, column=20, value=renewal.status)
        ws.cell(row=row, column=21, value=str(renewal.reminder_due) if renewal.reminder_due else None)
        ws.cell(row=row, column=22, value='Yes' if renewal.data_confirmed else 'No')
        ws.cell(row=row, column=23, value=renewal.last_known_year)
        ws.cell(row=row, column=24, value=renewal.notes)
        ws.cell(row=row, column=25, value=renewal.employee.name if renewal.employee else None)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"BE_RTO_Export_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)


@employees.route('/employees')
@login_required
def list_employees():
    if current_user.role != 'admin':
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard.index'))
    all_employees = Employee.query.order_by(Employee.name).all()
    return render_template('employees/list.html', employees=all_employees)

@employees.route('/employees/new', methods=['GET', 'POST'])
@login_required
def new_employee():
    if current_user.role != 'admin':
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard.index'))

    if request.method == 'POST':
        employee = Employee(
            name=request.form.get('name'),
            email=request.form.get('email'),
            branch=request.form.get('branch'),
            role=request.form.get('role')
        )
        employee.set_password(request.form.get('password'))
        db.session.add(employee)
        db.session.commit()
        flash('Employee added.', 'success')
        return redirect(url_for('employees.list_employees'))

    return render_template('employees/form.html')

@employees.route('/employees/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_employee(id):
    if current_user.role != 'admin':
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard.index'))

    employee = Employee.query.get_or_404(id)

    if request.method == 'POST':
        employee.name = request.form.get('name')
        employee.email = request.form.get('email')
        employee.branch = request.form.get('branch')
        employee.role = request.form.get('role')

        new_password = request.form.get('password')
        if new_password:
            employee.set_password(new_password)

        db.session.commit()
        flash('Employee updated.', 'success')
        return redirect(url_for('employees.list_employees'))

    return render_template('employees/edit.html', employee=employee)