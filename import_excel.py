import openpyxl
from app import create_app, db
from app.models import Customer, Vehicle, Renewal, Insurer
from datetime import datetime, timedelta

app = create_app()

INSURER_MAP = {
    'bajaj': 'Bajaj Allianz',
    'digit': 'Go Digit',
    'go digit': 'Go Digit',
    'iffco': 'IFFCO-Tokio',
    'reliance': 'Reliance General',
    'united': 'United India',
    'national': 'National Insurance',
    'chola': 'Cholamandalam',
    'kotak': 'Kotak General',
    'new india': 'New India Assurance',
    'zuno': 'Zuno',
}

def match_insurer(name):
    if not name:
        return None
    name_lower = str(name).lower()
    for key, val in INSURER_MAP.items():
        if key in name_lower:
            return val
    return str(name)

def parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    try:
        return datetime.strptime(str(val), '%Y-%m-%d').date()
    except:
        try:
            return datetime.strptime(str(val), '%d/%m/%Y').date()
        except:
            return None

with app.app_context():
    wb = openpyxl.load_workbook('Insurance policy OLD.xlsx', data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\nProcessing sheet: {sheet_name}')
        rows = list(ws.iter_rows(values_only=True))

        header = rows[0]
        print(f'Columns: {header}')

        for row in rows[1:]:
            if not any(row):
                continue

            try:
                customer_name = str(row[1]).strip() if row[1] else None

                SKIP_VALUES = {
                'customer name', 'inv.no.', 'sr no.', 'sr no', 'none', 
                'name', 'total payable amt.', 'expir date', ''
                }
                if not customer_name or customer_name.lower() in SKIP_VALUES:
                    continue

                policy_name = str(row[4]).strip() if row[4] else None
                policy_number = str(row[5]).strip() if row[5] else None
                start_date = parse_date(row[6])
                expiry_date = parse_date(row[7])
                model = str(row[9]).strip() if row[9] else 'Unknown'
                chassis_no = str(row[10]).strip() if row[10] else None
                engine_no = str(row[11]).strip() if row[11] else None

                customer = Customer.query.filter_by(name=customer_name).first()
                if not customer:
                    customer = Customer(
                        name=customer_name,
                        branch='Navi Mumbai',
                        is_company=False
                    )
                    db.session.add(customer)
                    db.session.flush()

                vehicle = Vehicle(
                    customer_id=customer.id,
                    model=model,
                    category='Other',
                    chassis_no=chassis_no,
                    engine_no=engine_no
                )
                db.session.add(vehicle)
                db.session.flush()

                insurer_name = match_insurer(policy_name)
                insurer = None
                if insurer_name:
                    insurer = Insurer.query.filter_by(name=insurer_name).first()

                reminder_due = expiry_date - timedelta(days=60) if expiry_date else None

                renewal = Renewal(
                    vehicle_id=vehicle.id,
                    insurer_id=insurer.id if insurer else None,
                    policy_number=policy_number,
                    policy_start=start_date,
                    policy_end=expiry_date,
                    reminder_due=reminder_due,
                    status='open',
                    data_confirmed=True
                )
                db.session.add(renewal)

            except Exception as e:
                print(f'Error on row {row}: {e}')
                continue

        db.session.commit()
        print(f'Sheet {sheet_name} imported.')

    print('\nImport complete.')